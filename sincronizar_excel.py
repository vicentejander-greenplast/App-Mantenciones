# =============================================================
#  GREENPLAST — Sincronizador de Excel Local
#
#  Este script corre en un PC de la planta con acceso a G:\
#  Descarga todos los registros desde la app en Railway
#  y actualiza la Planilla Maestra de Mantenciones.
#
#  Ejecución automática: configurar en el Programador de
#  Tareas de Windows (ver sincronizar.bat)
# =============================================================

import os
import sys
import json
import urllib.request
from datetime import datetime

# ── CONFIGURACIÓN ─────────────────────────────────────────────
RAILWAY_URL = "https://app-mantenciones-production.up.railway.app"
API_SECRET  = "gp-cal-sync-2024"
EXCEL_PATH  = r"G:\Unidades compartidas\GREENPLAST\Mantención\Maestro\Planilla_Maestro_Mantenciones.xlsx"

# ── DEPENDENCIAS ──────────────────────────────────────────────
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instala openpyxl: pip install openpyxl")
    sys.exit(1)

# ── COLORES GREENPLAST ────────────────────────────────────────
COLOR_VERDE_OSCURO = "1B5E20"
COLOR_VERDE_MEDIO  = "388E3C"
COLOR_VERDE_SUAVE  = "E8F5E9"
COLOR_BLANCO       = "FFFFFF"

HEADERS = [
    "ID", "Fecha Registro", "Fecha", "Hora", "Máquina",
    "Tipo", "Motivo", "Descripción", "Técnico",
    "Duración (hrs)", "Estado", "Programada", "Observaciones",
]
WIDTHS = [6, 18, 12, 8, 22, 18, 28, 38, 22, 12, 14, 12, 30]


def fetch_records():
    """Descarga todos los registros desde Railway."""
    url = f"{RAILWAY_URL}/api/todos?secret={API_SECRET}"
    try:
        with urllib.request.urlopen(url, timeout=15) as resp:
            return json.loads(resp.read().decode())
    except Exception as e:
        print(f"[ERROR] No se pudo conectar con Railway: {e}")
        return None


def get_existing_ids(ws):
    """Retorna los IDs que ya están en el Excel."""
    ids = set()
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is not None:
            try:
                ids.add(int(row[0]))
            except (ValueError, TypeError):
                pass
    return ids


def create_headers(ws):
    """Crea encabezados con estilo Greenplast."""
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:M1")
    ws["A1"] = "🔧  GREENPLAST — PLANILLA MAESTRA DE MANTENCIONES"
    ws["A1"].font = Font(bold=True, size=14, color=COLOR_BLANCO)
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_VERDE_OSCURO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    ws["A2"] = "Planta de Reciclaje HDPE y PP Rígido — Greenplast"
    ws["A2"].font = Font(italic=True, size=10, color=COLOR_BLANCO)
    ws["A2"].fill = PatternFill("solid", fgColor=COLOR_VERDE_MEDIO)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for col_i, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col_i, value=header)
        cell.font = Font(bold=True, size=10, color=COLOR_BLANCO)
        cell.fill = PatternFill("solid", fgColor=COLOR_VERDE_MEDIO)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 36

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"


def append_record(ws, rec, row_i):
    """Agrega una fila de registro al Excel."""
    bg = COLOR_VERDE_SUAVE if row_i % 2 == 0 else COLOR_BLANCO
    fill = PatternFill("solid", fgColor=bg)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    values = [
        rec.get("id"),
        rec.get("fecha_registro"),
        rec.get("fecha"),
        rec.get("hora"),
        rec.get("maquina"),
        rec.get("tipo"),
        rec.get("motivo"),
        rec.get("descripcion"),
        rec.get("tecnico"),
        rec.get("duracion"),
        rec.get("estado"),
        rec.get("programada"),
        rec.get("observaciones"),
    ]
    for col_i, val in enumerate(values, 1):
        cell = ws.cell(row=row_i, column=col_i, value=val)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row_i].height = 20


def sync():
    print(f"\n[{datetime.now().strftime('%d/%m/%Y %H:%M')}] Iniciando sincronización...")

    # Descargar registros desde Railway
    records = fetch_records()
    if records is None:
        print("Sincronización cancelada por error de conexión.")
        return

    print(f"  {len(records)} registros en la app.")

    # Cargar o crear Excel
    if os.path.exists(EXCEL_PATH):
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        existing_ids = get_existing_ids(ws)
        print(f"  {len(existing_ids)} registros ya en el Excel.")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Mantenciones"
        create_headers(ws)
        existing_ids = set()
        print("  Planilla nueva creada.")

    # Agregar solo registros nuevos
    new_records = [r for r in records if r.get("id") not in existing_ids]
    if not new_records:
        print("  Sin registros nuevos. Excel actualizado.")
        return

    next_row = ws.max_row + 1
    for rec in new_records:
        append_record(ws, rec, next_row)
        next_row += 1

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"  ✅ {len(new_records)} registros nuevos agregados al Excel.")
    print(f"  Guardado en: {EXCEL_PATH}")


if __name__ == "__main__":
    sync()
