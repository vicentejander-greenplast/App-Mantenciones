# =============================================================
#  GREENPLAST - Generador de Excel bajo demanda
#  Se genera desde la BD SQLite cuando el usuario descarga.
# =============================================================

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLOR_VERDE_OSCURO = "1B5E20"
COLOR_VERDE_MEDIO  = "388E3C"
COLOR_VERDE_SUAVE  = "E8F5E9"
COLOR_BLANCO       = "FFFFFF"

HEADERS = [
    "ID", "Fecha Registro", "Fecha", "Hora", "Máquina",
    "Tipo", "Motivo", "Descripción", "Técnico",
    "Duración (hrs)", "Estado", "Programada", "Observaciones",
]
WIDTHS = [6, 18, 12, 8, 22, 18, 28, 38, 22, 12, 14, 12, 30]


def generate_excel_bytes(records: list) -> bytes:
    """
    Genera un archivo Excel en memoria desde una lista de registros
    y retorna los bytes para enviar como descarga.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mantenciones"

    # ── Fila 1: Título ──────────────────────────────────────────
    ws.merge_cells("A1:M1")
    ws["A1"] = "🔧  GREENPLAST — PLANILLA MAESTRA DE MANTENCIONES"
    ws["A1"].font      = Font(bold=True, size=14, color=COLOR_BLANCO)
    ws["A1"].fill      = PatternFill("solid", fgColor=COLOR_VERDE_OSCURO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:M2")
    ws["A2"] = f"Exportado el {datetime.now().strftime('%d/%m/%Y %H:%M')} — Planta de Reciclaje HDPE y PP"
    ws["A2"].font      = Font(italic=True, size=10, color=COLOR_BLANCO)
    ws["A2"].fill      = PatternFill("solid", fgColor=COLOR_VERDE_MEDIO)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Fila 3: Encabezados ─────────────────────────────────────
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col_i, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=3, column=col_i, value=header)
        cell.font      = Font(bold=True, size=10, color=COLOR_BLANCO)
        cell.fill      = PatternFill("solid", fgColor=COLOR_VERDE_MEDIO)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[3].height = 36
    ws.freeze_panes = "A4"

    # ── Anchos ──────────────────────────────────────────────────
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Datos ────────────────────────────────────────────────────
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for row_i, rec in enumerate(records, start=4):
        bg = COLOR_VERDE_SUAVE if row_i % 2 == 0 else COLOR_BLANCO
        fill = PatternFill("solid", fgColor=bg)
        values = [
            rec.get("id"),
            rec.get("fecha_registro"),
            rec.get("fecha"),
            rec.get("hora"),
            rec.get("maquina"),
            rec.get("tipo"),
            rec.get("motivo"),
            rec.get("descripcion"),
            rec.get("tecnico"),
            rec.get("duracion"),
            rec.get("estado"),
            rec.get("programada"),
            rec.get("observaciones"),
        ]
        for col_i, val in enumerate(values, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row_i].height = 20

    # ── Guardar en memoria ───────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
